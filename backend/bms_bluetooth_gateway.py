#!/usr/bin/env python3
"""
BMS Bluetooth Gateway
---------------------
This script runs on the Raspberry Pi. It connects to a physical JBD (Xiaoxiang) BMS
or Daly BMS via BLE, parses telemetry data using your exact logic, and forwards it to the 
Flask backend's telemetry endpoint (/api/telemetry).
"""

import os
import sys
import time
import json
import asyncio
import argparse
import urllib.request
from datetime import datetime
from typing import Dict, Any, Optional

try:
    from bleak import BleakClient, BleakScanner
except ImportError:
    BleakClient = None
    BleakScanner = None

# ==============================================================================
# DEFAULT CONFIGURATION (Defaults to your exact JBD BMS hardware)
# ==============================================================================
DEFAULT_BMS_TYPE = "JBD"
DEFAULT_MAC_ADDRESS = "A4:C1:37:04:28:FB"  # Your JBD BMS MAC Address
DEFAULT_API_URL = "http://127.0.0.1:5000/api/telemetry"
DEFAULT_POLL_INTERVAL = 1.0
# ==============================================================================

# JBD BLE Constants
JBD_NOTIFY_UUID = "0000ff01-0000-1000-8000-00805f9b34fb"
JBD_WRITE_UUID  = "0000ff02-0000-1000-8000-00805f9b34fb"

JBD_CMD_BASIC = bytes([0xDD, 0xA5, 0x03, 0x00, 0xFF, 0xFD, 0x77])
JBD_CMD_CELL  = bytes([0xDD, 0xA5, 0x04, 0x00, 0xFF, 0xFC, 0x77])

V_MIN = 22.4
V_MAX = 33.6

# Daly BLE Constants
DALY_SERVICE_UUID = "6e400001-b5a3-f393-e0a9-e50e24dcca9e"
DALY_TX_UUID = "6e400002-b5a3-f393-e0a9-e50e24dcca9e"
DALY_RX_UUID = "6e400003-b5a3-f393-e0a9-e50e24dcca9e"

def make_daly_cmd(cmd_id: int) -> bytes:
    payload = bytes([0xA5, 0x40, cmd_id, 0x08, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00])
    checksum = sum(payload) & 0xFF
    return payload + bytes([checksum])


# ==============================================================================
# JBD PARSING LOGIC (EXACTLY MATCHING YOUR WORKING VERSION)
# ==============================================================================
def parse_jbd_temperature(data):
    try:
        # data[26] contains the NTC count. The raw temperature payload starts at index 27.
        if len(data) < 27:
            return None, None, None, None

        ntc_count = data[26]
        temps = []
        for i in range(4):
            if i < ntc_count:
                idx = 27 + i * 2
                if idx + 2 <= len(data):
                    val = int.from_bytes(data[idx:idx + 2], "big")
                    t = round((val - 2731) / 10, 1)
                    temps.append(t if -40 <= t <= 120 else None)
                else:
                    temps.append(None)
            else:
                temps.append(None)
        return temps[0], temps[1], temps[2], temps[3]
    except Exception:
        return None, None, None, None

def parse_jbd_basic(data):
    try:
        voltage = int.from_bytes(data[4:6], "big") / 100.0

        raw_i = int.from_bytes(data[6:8], "big", signed=False)
        if raw_i & 0x8000:
            raw_i -= 0x10000
        current = raw_i / 100.0

        # SOC parsing: Extract the real SOC from the BMS Coulomb counter (byte 23) if available
        soc = float(data[23]) if len(data) >= 24 else round((voltage - V_MIN) / (V_MAX - V_MIN) * 100, 1)
        soc = max(0.0, min(100.0, soc))

        ntc1, ntc2, ntc3, ntc4 = parse_jbd_temperature(data)
        valid_temps = [t for t in (ntc1, ntc2, ntc3, ntc4) if t is not None]
        temperature = max(valid_temps) if valid_temps else None

        return voltage, current, soc, temperature, ntc1, ntc2, ntc3, ntc4
    except Exception:
        return 0.0, 0.0, 0.0, None, None, None, None, None

def parse_jbd_cells(data):
    cells = []
    try:
        # data[3] is the length byte L. Number of cells is L // 2.
        L = data[3]
        num_cells = L // 2
        for i in range(num_cells):
            idx = 4 + i * 2
            if idx + 2 <= len(data):
                v = int.from_bytes(data[idx:idx + 2], "big") / 1000
                if 0.5 <= v <= 5.0:
                    cells.append(round(v, 3))
                else:
                    cells.append(None)
            else:
                cells.append(None)
    except Exception:
        cells = [3.2] * 8

    # Extract only the actual connected cell channels (voltage >= 0.5V)
    # This prevents the empty channels (which report 0.001V) from shifting the cell numbers
    valid_cells = [c for c in cells if c is not None and c >= 0.5]
    
    # Calculate the average of the connected cells
    mean_v = round(sum(valid_cells) / len(valid_cells), 3) if valid_cells else 3.2
    
    # Round the valid cell voltages
    cleaned_cells = [round(c, 3) for c in valid_cells]
    
    # Pad at the end of the list up to 8 cells for backend validation
    while len(cleaned_cells) < 8:
        cleaned_cells.append(mean_v)
        
    return cleaned_cells


# ==============================================================================
# GATEWAY CLASS
# ==============================================================================
class BMSGateway:
    def __init__(self, bms_type: str, mac_address: str, api_url: str, poll_interval: float, verbose: bool):
        self.bms_type = bms_type.upper()
        self.mac_address = mac_address
        self.api_url = api_url
        self.poll_interval = poll_interval
        self.verbose = verbose
        
        # Debugging and validity tracking fields
        self.packet_count = 0
        self.parse_error_count = 0
        self.last_parse_error = "None"
        self.reconnect_attempt_count = 0
        self.last_successful_packet_time = 0.0
        self.raw_packet_hex = "N/A"
        self.current_raw_value = 0
        self.current_scaled_value = 0.0
        self.bluetooth_connected = False
        
        # Responses for notifications
        self.response_basic = None
        self.response_cells = None
        self.rx_buffer = bytearray()
        
        # Telemetry data buffer
        self.latest_data: Dict[str, Any] = {}
        for i in range(1, 9):
            self.latest_data[f"cell_v{i}"] = 3.2
        self.latest_data.update({
            "voltage": 25.6,
            "current": 0.0,
            "temperature": 25.0,
            "soc": 50.0,
            "delta_v": 0.0,
            "ntc1": 25.0,
            "ntc2": 25.0,
            "ntc3": 25.0,
            "ntc4": 25.0
        })

    def log(self, msg: str):
        print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {msg}")

    def log_verbose(self, msg: str):
        if self.verbose:
            self.log(f"[DEBUG] {msg}")

    def log_to_local_csv(self):
        """Append the latest telemetry frame directly to a local CSV file."""
        csv_file = "bms_local_telemetry_log.csv"
        file_exists = os.path.exists(csv_file)
        
        headers = [
            "timestamp", "voltage", "current", "temperature", "soc", "delta_v", 
            "cell_v1", "cell_v2", "cell_v3", "cell_v4", "cell_v5", "cell_v6", "cell_v7", "cell_v8",
            "ntc1", "ntc2", "ntc3", "ntc4", "bluetooth_connected"
        ]
        
        row = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
            "voltage": self.latest_data.get("voltage", 0.0),
            "current": self.latest_data.get("current", 0.0),
            "temperature": self.latest_data.get("temperature", 0.0),
            "soc": self.latest_data.get("soc", 0.0),
            "delta_v": self.latest_data.get("delta_v", 0.0),
            "cell_v1": self.latest_data.get("cell_v1", 3.2),
            "cell_v2": self.latest_data.get("cell_v2", 3.2),
            "cell_v3": self.latest_data.get("cell_v3", 3.2),
            "cell_v4": self.latest_data.get("cell_v4", 3.2),
            "cell_v5": self.latest_data.get("cell_v5", 3.2),
            "cell_v6": self.latest_data.get("cell_v6", 3.2),
            "cell_v7": self.latest_data.get("cell_v7", 3.2),
            "cell_v8": self.latest_data.get("cell_v8", 3.2),
            "ntc1": self.latest_data.get("ntc1", 25.0),
            "ntc2": self.latest_data.get("ntc2", 25.0),
            "ntc3": self.latest_data.get("ntc3", 25.0),
            "ntc4": self.latest_data.get("ntc4", 25.0),
            "bluetooth_connected": self.latest_data.get("bluetooth_connected", False)
        }
        
        try:
            import csv
            with open(csv_file, mode="a", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                if not file_exists:
                    writer.writeheader()
                writer.writerow(row)
        except Exception as e:
            self.log(f"Error writing to local CSV: {e}")

    def log_to_local_xlsx(self):
        """Append the latest telemetry frame directly to the local Excel file in Downloads."""
        xlsx_file = r"C:\Users\aksha\Downloads\bms_local_telemetry_log.xlsx"
        
        headers = [
            "timestamp", "voltage", "current", "temperature", "soc", "delta_v", 
            "cell_v1", "cell_v2", "cell_v3", "cell_v4", "cell_v5", "cell_v6", "cell_v7", "cell_v8",
            "ntc1", "ntc2", "ntc3", "ntc4", "bluetooth_connected"
        ]
        
        row_values = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
            self.latest_data.get("voltage", 0.0),
            self.latest_data.get("current", 0.0),
            self.latest_data.get("temperature", 0.0),
            self.latest_data.get("soc", 0.0),
            self.latest_data.get("delta_v", 0.0),
            self.latest_data.get("cell_v1", 3.2),
            self.latest_data.get("cell_v2", 3.2),
            self.latest_data.get("cell_v3", 3.2),
            self.latest_data.get("cell_v4", 3.2),
            self.latest_data.get("cell_v5", 3.2),
            self.latest_data.get("cell_v6", 3.2),
            self.latest_data.get("cell_v7", 3.2),
            self.latest_data.get("cell_v8", 3.2),
            self.latest_data.get("ntc1", 25.0),
            self.latest_data.get("ntc2", 25.0),
            self.latest_data.get("ntc3", 25.0),
            self.latest_data.get("ntc4", 25.0),
            self.latest_data.get("bluetooth_connected", False)
        ]
        
        try:
            import openpyxl
            try:
                if not os.path.exists(xlsx_file):
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Telemetry Log"
                    ws.append(headers)
                    wb.save(xlsx_file)
                wb = openpyxl.load_workbook(xlsx_file)
            except Exception:
                # If file exists but is corrupted, delete and recreate it
                if os.path.exists(xlsx_file):
                    try:
                        os.remove(xlsx_file)
                    except Exception:
                        pass
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Telemetry Log"
                ws.append(headers)
                wb.save(xlsx_file)
                wb = openpyxl.load_workbook(xlsx_file)
                
            ws = wb.active
            ws.append(row_values)
            wb.save(xlsx_file)
        except Exception as e:
            # Excel might be open/locked by the user, log it but don't crash
            self.log(f"Error writing to local Excel file: {e}")

    def post_telemetry(self) -> bool:
        """POSTs current telemetry payload to the Flask backend and logs locally."""
        # Save telemetry to local CSV and Excel files
        self.log_to_local_csv()
        self.log_to_local_xlsx()
        
        try:
            req_data = json.dumps(self.latest_data).encode("utf-8")
            headers = {
                "Content-Type": "application/json",
                "X-Ingest-Token": os.environ.get("INGEST_TOKEN", "")
            }
            req = urllib.request.Request(
                self.api_url,
                data=req_data,
                headers=headers
            )
            with urllib.request.urlopen(req, timeout=2.0) as resp:
                result = json.loads(resp.read().decode("utf-8"))
                status = result.get("status")
                buf_len = result.get("buffer_length", 0)
                pred = result.get("prediction")
                
                # Standard console logger output format
                volts = self.latest_data.get('voltage', 0.0)
                curr = self.latest_data.get('current', 0.0)
                soc = self.latest_data.get('soc', 0.0)
                if status == "success" and pred:
                    self.log(f"Row sent! Buffer: {buf_len}/60 | V={volts}V, I={curr}A, SOC={soc}% | Fault: {pred.get('Fault Prediction')} ({pred.get('Confidence Score')})")
                else:
                    self.log(f"Row sent! Buffer: {buf_len}/60 | V={volts}V, I={curr}A, SOC={soc}% (Accumulating window)")
                return True
        except Exception as e:
            self.log(f"Error POSTing to backend: {e}. Make sure Flask server is running on Port 5000!")
            return False

    # --------------------------------------------------------------------------
    # MOCK MODE
    # --------------------------------------------------------------------------
    async def run_mock(self):
        self.log("Starting in MOCK Mode. Generating synthetic BMS data...")
        import math
        t_counter = 0
        while True:
            t_counter += self.poll_interval
            current = round(5.0 * math.sin(t_counter / 30.0), 2)
            base_volts = 25.6
            voltage = round(base_volts - (current * 0.15) + 0.5 * math.sin(t_counter / 60.0), 2)
            temperature = round(25.0 + abs(current) * 0.4 + 2.0 * math.sin(t_counter / 120.0), 1)
            soc = round(max(0.0, min(100.0, 75.0 + 5.0 * math.sin(t_counter / 180.0))), 1)
            
            cell_v = []
            avg_cell = voltage / 8.0
            for i in range(8):
                offset = 0.02 * math.sin(t_counter / 10.0 + i)
                if i == 3 and t_counter > 90.0:
                    offset -= 0.15 * min(1.0, (t_counter - 90.0) / 60.0)
                cell_v.append(round(avg_cell + offset, 3))
                
            self.latest_data = {
                "voltage": voltage,
                "current": current,
                "temperature": temperature,
                "soc": soc,
                "delta_v": round(max(cell_v) - min(cell_v), 3)
            }
            for i in range(8):
                self.latest_data[f"cell_v{i+1}"] = cell_v[i]
            for n in range(1, 5):
                self.latest_data[f"ntc{n}"] = round(temperature + (n - 2.5) * 0.3, 1)
                
            self.post_telemetry()
            await asyncio.sleep(self.poll_interval)

    # --------------------------------------------------------------------------
    # JBD / XIAOXIANG BLE CLIENT
    # --------------------------------------------------------------------------
    async def run_jbd(self):
        if BleakClient is None:
            self.log("ERROR: bleak library not found. Run 'pip install bleak'")
            return

        self.log(f"Connecting to JBD BMS at {self.mac_address}...")
        async with BleakClient(self.mac_address, timeout=15.0) as client:
            self.log("Connected to JBD BMS! ✅")
            self.bluetooth_connected = True
            
            def handle_notify(sender, data):
                self.rx_buffer.extend(data)
                while len(self.rx_buffer) >= 7:
                    if 0xDD not in self.rx_buffer:
                        self.rx_buffer.clear()
                        break
                    start = self.rx_buffer.index(0xDD)
                    if start > 0:
                        self.rx_buffer = self.rx_buffer[start:]
                        start = 0
                        
                    if len(self.rx_buffer) < 4:
                        break
                        
                    cmd = self.rx_buffer[1]
                    status = self.rx_buffer[2]
                    L = self.rx_buffer[3]
                    
                    expected_len = L + 7
                    if len(self.rx_buffer) < expected_len:
                        break
                        
                    packet = bytes(self.rx_buffer[:expected_len])
                    
                    # Verify stop byte at correct offset
                    if packet[L + 6] != 0x77:
                        self.log_verbose(f"Invalid stop byte {packet[L + 6]:02X} (expected 0x77) for command {cmd:02X}. Discarding header.")
                        self.rx_buffer.pop(0)
                        self.parse_error_count += 1
                        self.last_parse_error = f"Invalid stop byte {packet[L + 6]:02X}"
                        continue
                        
                    # Verify checksum (excludes start byte 0xDD and command byte at packet[1])
                    total_sum = sum(packet[2 : 4 + L])
                    checksum_calc = (0x10000 - total_sum) & 0xFFFF
                    checksum_packet = int.from_bytes(packet[4 + L : 4 + L + 2], "big")
                    
                    if checksum_calc != checksum_packet:
                        self.log(f"Checksum mismatch for command {cmd:02X}: calculated {checksum_calc:04X}, packet {checksum_packet:04X}. Discarding packet.")
                        self.rx_buffer = self.rx_buffer[expected_len:]
                        self.parse_error_count += 1
                        self.last_parse_error = f"Checksum mismatch for command {cmd:02X}"
                        continue
                        
                    # Valid packet! Remove from buffer
                    self.rx_buffer = self.rx_buffer[expected_len:]
                    self.packet_count += 1
                    self.last_successful_packet_time = time.time()
                    self.raw_packet_hex = packet.hex()
                    
                    if cmd == 0x03:
                        self.response_basic = packet
                    elif cmd == 0x04:
                        self.response_cells = packet

            await client.start_notify(JBD_NOTIFY_UUID, handle_notify)
            
            while True:
                # Query Basic Information
                await client.write_gatt_char(JBD_WRITE_UUID, JBD_CMD_BASIC)
                await asyncio.sleep(0.15)
                
                # Query Cell Voltages
                await client.write_gatt_char(JBD_WRITE_UUID, JBD_CMD_CELL)
                await asyncio.sleep(0.15)
                
                # If we have responses, parse and update
                if self.response_basic and self.response_cells:
                    v, i, soc, t, ntc1, ntc2, ntc3, ntc4 = parse_jbd_basic(self.response_basic)
                    cells = parse_jbd_cells(self.response_cells)
                    # Keep the true total pack voltage from basic info instead of summing padded cells
                    delta_v = round(max(cells) - min(cells), 4)
                    
                    # Print raw packet debug output
                    print("--------------------------------------------------")
                    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] JBD BMS DEBUG PACKET RECEIVED")
                    print(f"  Basic Raw: {self.response_basic.hex().upper()} (Len: {len(self.response_basic)})")
                    print(f"  Cells Raw: {self.response_cells.hex().upper()} (Len: {len(self.response_cells)})")
                    print(f"  Parsed Voltage: {v} V")
                    print(f"  Parsed Current: {i} A (Raw: {int.from_bytes(self.response_basic[6:8], 'big', signed=True)})")
                    print(f"  Parsed Temp: {t} °C")
                    print(f"  Parsed Cells: {cells}")
                    print("--------------------------------------------------")
                    
                    # Check physical limits for validity flags
                    current_valid = (-120.0 <= i <= 120.0)
                    voltage_valid = (15.0 <= v <= 40.0)
                    temperature_valid = (t is None or -40.0 <= t <= 120.0)
                    cell_voltage_valid = all(0.5 <= cv <= 5.0 for cv in cells)
                    
                    self.latest_data = {
                        "voltage": v,
                        "current": i,
                        "temperature": t,
                        "ntc1": ntc1,
                        "ntc2": ntc2,
                        "ntc3": ntc3,
                        "ntc4": ntc4,
                        "soc": soc,
                        "delta_v": delta_v,
                        "bluetooth_connected": True,
                        "reconnect_attempt_count": self.reconnect_attempt_count,
                        "last_successful_packet_time": self.last_successful_packet_time,
                        "raw_packet_hex": self.response_basic.hex(),
                        "packet_count": self.packet_count,
                        "parse_error_count": self.parse_error_count,
                        "last_parse_error": self.last_parse_error,
                        "current_raw_value": int.from_bytes(self.response_basic[6:8], 'big', signed=True),
                        "current_scaled_value": i,
                        "current_valid": current_valid,
                        "voltage_valid": voltage_valid,
                        "temperature_valid": temperature_valid,
                        "cell_voltage_valid": cell_voltage_valid
                    }
                    for idx in range(8):
                        self.latest_data[f"cell_v{idx+1}"] = cells[idx]
                        
                    # Send payload to server
                    self.post_telemetry()
                    
                await asyncio.sleep(max(0.1, self.poll_interval - 0.3))

    # --------------------------------------------------------------------------
    # DALY BLE CLIENT (Nordic NUS transparent serial fallback)
    # --------------------------------------------------------------------------
    def parse_daly_notification(self, sender: int, data: bytearray):
        self.rx_buffer.extend(data)
        while len(self.rx_buffer) >= 13:
            if self.rx_buffer[0] != 0xA5:
                self.rx_buffer.pop(0)
                continue
                
            packet = self.rx_buffer[:13]
            self.rx_buffer = self.rx_buffer[13:]
            
            checksum = sum(packet[:12]) & 0xFF
            if checksum != packet[12]:
                self.parse_error_count += 1
                self.last_parse_error = "Daly checksum mismatch"
                continue
                
            self.packet_count += 1
            self.last_successful_packet_time = time.time()
            self.raw_packet_hex = packet.hex()
            
            cmd = packet[2]
            payload = packet[4:12]
            
            if cmd == 0x90:
                voltage = (payload[0] << 8 | payload[1]) * 0.1
                raw_curr = payload[4] << 8 | payload[5]
                current = - ((raw_curr - 30000) * 0.1)
                soc = (payload[6] << 8 | payload[7]) * 0.1
                self.latest_data["voltage"] = round(voltage, 2)
                self.latest_data["current"] = round(current, 2)
                self.latest_data["soc"] = round(soc, 1)
                
                self.current_raw_value = raw_curr
                self.current_scaled_value = round(current, 2)
                
                print("--------------------------------------------------")
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] DALY BMS DEBUG PACKET (0x90)")
                print(f"  Raw Frame: {packet.hex().upper()}")
                print(f"  Parsed Voltage: {voltage} V")
                print(f"  Parsed Current: {current} A (Raw: {raw_curr})")
                print(f"  Parsed SOC: {soc} %")
                print("--------------------------------------------------")
            elif cmd == 0x91:
                max_v = (payload[0] << 8 | payload[1]) * 0.001
                min_v = (payload[3] << 8 | payload[4]) * 0.001
                self.latest_data["delta_v"] = round(max_v - min_v, 3)
            elif cmd == 0x92:
                max_t = payload[0] - 40
                min_t = payload[2] - 40
                avg_t = (max_t + min_t) / 2.0
                self.latest_data["temperature"] = round(avg_t, 1)
                self.latest_data["ntc1"] = float(max_t)
                self.latest_data["ntc2"] = float(min_t)
                self.latest_data["ntc3"] = float(avg_t)
                self.latest_data["ntc4"] = float(avg_t)
            elif cmd == 0x95:
                frame_idx = payload[0]
                cell_start = (frame_idx - 1) * 3
                for offset in range(3):
                    cell_num = cell_start + offset + 1
                    if cell_num <= 8:
                        byte_offset = 1 + offset * 2
                        val = payload[byte_offset] << 8 | payload[byte_offset+1]
                        self.latest_data[f"cell_v{cell_num}"] = round(val * 0.001, 3)

    async def run_daly(self):
        if BleakClient is None:
            self.log("ERROR: bleak library not found. Run 'pip install bleak'")
            return

        self.log(f"Connecting to Daly BMS at {self.mac_address}...")
        async with BleakClient(self.mac_address, timeout=15.0) as client:
            self.log("Connected to Daly BMS! ✅")
            self.bluetooth_connected = True
            await client.start_notify(DALY_RX_UUID, self.parse_daly_notification)
            while True:
                await client.write_gatt_char(DALY_TX_UUID, make_daly_cmd(0x90), response=False)
                await asyncio.sleep(0.2)
                await client.write_gatt_char(DALY_TX_UUID, make_daly_cmd(0x91), response=False)
                await asyncio.sleep(0.2)
                await client.write_gatt_char(DALY_TX_UUID, make_daly_cmd(0x92), response=False)
                await asyncio.sleep(0.2)
                await client.write_gatt_char(DALY_TX_UUID, make_daly_cmd(0x95), response=False)
                await asyncio.sleep(0.4)
                
                v = self.latest_data.get("voltage", 25.6)
                i = self.latest_data.get("current", 0.0)
                t = self.latest_data.get("temperature", 25.0)
                cells = [self.latest_data.get(f"cell_v{idx}", 3.2) for idx in range(1, 9)]
                
                current_valid = (-120.0 <= i <= 120.0)
                voltage_valid = (15.0 <= v <= 40.0)
                temperature_valid = (-40.0 <= t <= 120.0)
                cell_voltage_valid = all(2.0 <= cv <= 4.6 for cv in cells)
                
                self.latest_data.update({
                    "bluetooth_connected": True,
                    "reconnect_attempt_count": self.reconnect_attempt_count,
                    "last_successful_packet_time": self.last_successful_packet_time,
                    "raw_packet_hex": self.raw_packet_hex,
                    "packet_count": self.packet_count,
                    "parse_error_count": self.parse_error_count,
                    "last_parse_error": self.last_parse_error,
                    "current_raw_value": self.current_raw_value,
                    "current_scaled_value": self.current_scaled_value,
                    "current_valid": current_valid,
                    "voltage_valid": voltage_valid,
                    "temperature_valid": temperature_valid,
                    "cell_voltage_valid": cell_voltage_valid
                })
                
                self.post_telemetry()
                await asyncio.sleep(max(0.1, self.poll_interval - 1.0))

    # --------------------------------------------------------------------------
    # MAIN START LOOP WITH RECONNECTION HANDLER
    # --------------------------------------------------------------------------
    async def start(self):
        if self.bms_type == "MOCK":
            await self.run_mock()
            return
            
        while True:
            try:
                if self.bms_type == "JBD":
                    await self.run_jbd()
                elif self.bms_type == "DALY":
                    await self.run_daly()
                else:
                    self.log(f"Unknown BMS brand: {self.bms_type}. Defaulting to MOCK.")
                    await self.run_mock()
            except Exception as e:
                self.bluetooth_connected = False
                self.reconnect_attempt_count += 1
                self.last_parse_error = str(e)
                import traceback
                traceback.print_exc()
                self.log(f"Bluetooth connection error: {e}")
                self.log("Reconnecting in 5 seconds...")
                
                # Send offline status to backend immediately
                self.latest_data.update({
                    "bluetooth_connected": False,
                    "reconnect_attempt_count": self.reconnect_attempt_count,
                    "last_successful_packet_time": self.last_successful_packet_time,
                    "raw_packet_hex": "N/A",
                    "packet_count": self.packet_count,
                    "parse_error_count": self.parse_error_count,
                    "last_parse_error": self.last_parse_error,
                    "current_raw_value": 0,
                    "current_scaled_value": 0.0,
                    "current_valid": False,
                    "voltage_valid": False,
                    "temperature_valid": False,
                    "cell_voltage_valid": False
                })
                self.post_telemetry()
                
                self.rx_buffer.clear()
                self.response_basic = None
                self.response_cells = None
                await asyncio.sleep(5.0)


# ==============================================================================
# CLI EXECUTION
# ==============================================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="BMS Bluetooth Telemetry Gateway client.")
    parser.add_argument("--bms", type=str, default=DEFAULT_BMS_TYPE, choices=["MOCK", "JBD", "DALY"],
                        help="Select BMS Protocol brand (MOCK, JBD, or DALY)")
    parser.add_argument("--mac", type=str, default=DEFAULT_MAC_ADDRESS,
                        help="Bluetooth MAC address of the BMS")
    parser.add_argument("--url", type=str, default=DEFAULT_API_URL,
                        help="API Endpoint URL of the Flask server")
    parser.add_argument("--interval", type=float, default=DEFAULT_POLL_INTERVAL,
                        help="Polling interval in seconds")
    parser.add_argument("--verbose", action="store_true",
                        help="Print raw packet debug logs")
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("           BMS BLUETOOTH GATEWAY STARTING            ")
    print("=" * 60)
    print(f" BMS Model Protocol : {args.bms}")
    print(f" Target Device MAC   : {args.mac if args.bms != 'MOCK' else 'N/A'}")
    print(f" Local Flask API URL : {args.url}")
    print(f" Polling Interval    : {args.interval} seconds")
    print("=" * 60)
    
    gateway = BMSGateway(
        bms_type=args.bms,
        mac_address=args.mac,
        api_url=args.url,
        poll_interval=args.interval,
        verbose=args.verbose
    )
    
    try:
        asyncio.run(gateway.start())
    except KeyboardInterrupt:
        print("\nGateway stopped by user. Exiting.")
        sys.exit(0)
